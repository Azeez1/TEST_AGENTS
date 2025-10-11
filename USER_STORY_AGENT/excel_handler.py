"""
Excel Handler Module
Manages reading, writing, and appending to Excel files
"""

import os
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Handle Excel file operations for user stories"""

    COLUMNS = [
        "User Stories",
        "Feature/Epic",
        "Acceptance Criteria",
        "Rationale or Business Case",
        "Relevant Page(s)"
    ]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None

    def file_exists(self) -> bool:
        """Check if Excel file exists"""
        return os.path.exists(self.file_path)

    def load_workbook(self):
        """Load existing workbook or create new one"""
        if self.file_exists():
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self._create_header()

    def _create_header(self):
        """Create header row with formatting"""
        for col_num, column_name in enumerate(self.COLUMNS, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            # Header styling
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Set column widths
        self.worksheet.column_dimensions['A'].width = 50  # User Stories
        self.worksheet.column_dimensions['B'].width = 25  # Feature/Epic
        self.worksheet.column_dimensions['C'].width = 60  # Acceptance Criteria
        self.worksheet.column_dimensions['D'].width = 40  # Business Case
        self.worksheet.column_dimensions['E'].width = 25  # Relevant Pages

    def read_existing_stories(self) -> List[Dict]:
        """
        Read existing stories from the Excel file

        Returns:
            List of story dictionaries
        """
        if not self.file_exists():
            return []

        self.load_workbook()
        stories = []

        # Skip header row (row 1)
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If user story exists
                story = {
                    "user_story": row[0] or "",
                    "feature_epic": row[1] or "",
                    "acceptance_criteria": self._parse_ac_cell(row[2]) if row[2] else [],
                    "business_case": row[3] or "",
                    "relevant_pages": row[4] or ""
                }
                stories.append(story)

        return stories

    def _parse_ac_cell(self, ac_text: str) -> List[str]:
        """
        Parse acceptance criteria from a cell into a list

        Args:
            ac_text: Text from the AC cell

        Returns:
            List of individual acceptance criteria
        """
        if not ac_text:
            return []

        # Split by newlines and filter out empty lines
        acs = [line.strip().lstrip('- ').strip() for line in ac_text.split('\n') if line.strip()]
        return acs

    def write_stories(self, stories: List[Dict], append: bool = True):
        """
        Write stories to Excel file

        Args:
            stories: List of story dictionaries
            append: Whether to append to existing file or overwrite
        """
        if append and self.file_exists():
            self.load_workbook()
            start_row = self.worksheet.max_row + 1
        else:
            # Create new workbook (don't try to load empty/non-existent file)
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self._create_header()
            start_row = 2  # Start after header

        for story in stories:
            self._write_story_row(start_row, story)
            start_row += 1

        self.save()

    def _write_story_row(self, row_num: int, story: Dict):
        """
        Write a single story to a row

        Args:
            row_num: Row number to write to
            story: Story dictionary
        """
        # Format acceptance criteria with bullets and newlines
        ac_text = self._format_acceptance_criteria(story.get('acceptance_criteria', []))

        # Write data
        self.worksheet.cell(row=row_num, column=1).value = story.get('user_story', '')
        self.worksheet.cell(row=row_num, column=2).value = story.get('feature_epic', '')
        self.worksheet.cell(row=row_num, column=3).value = ac_text
        self.worksheet.cell(row=row_num, column=4).value = story.get('business_case', '')
        self.worksheet.cell(row=row_num, column=5).value = story.get('relevant_pages', '')

        # Apply formatting
        for col in range(1, 6):
            cell = self.worksheet.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Set row height to auto-fit content
        self.worksheet.row_dimensions[row_num].height = None

    def _format_acceptance_criteria(self, criteria: List[str]) -> str:
        """
        Format acceptance criteria list into a single cell string

        Args:
            criteria: List of AC strings

        Returns:
            Formatted string with bullets and newlines
        """
        if not criteria:
            return ""

        # Join with newlines and add bullet points
        return "\n".join(f"- {ac}" for ac in criteria)

    def update_story(self, row_num: int, story: Dict):
        """
        Update a specific story row

        Args:
            row_num: Row number to update (1-indexed, header is row 1)
            story: Updated story dictionary
        """
        self.load_workbook()
        # Add 1 because row 1 is header, so story 1 is row 2
        actual_row = row_num + 1
        self._write_story_row(actual_row, story)
        self.save()

    def save(self):
        """Save the workbook"""
        self.workbook.save(self.file_path)

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()


def create_excel_output(file_path: str, stories: List[Dict], append: bool = True):
    """
    Helper function to create/update Excel file with stories

    Args:
        file_path: Path to Excel file
        stories: List of story dictionaries
        append: Whether to append or overwrite
    """
    handler = ExcelHandler(file_path)
    handler.write_stories(stories, append=append)
    handler.close()


def read_existing_stories(file_path: str) -> List[Dict]:
    """
    Helper function to read existing stories from Excel

    Args:
        file_path: Path to Excel file

    Returns:
        List of story dictionaries
    """
    handler = ExcelHandler(file_path)
    stories = handler.read_existing_stories()
    handler.close()
    return stories
