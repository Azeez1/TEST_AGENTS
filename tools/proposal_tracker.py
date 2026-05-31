"""
Proposal Tracker — scans PROPOSAL_TEAM/outputs/<topic_id>/ folders and emits a
master Excel at PROPOSAL_TEAM/outputs/PROPOSAL_TRACKER.xlsx.

Sources scanned per topic:
  - PARTNER_CHECKLIST.md       — checkbox items + sign-off table
  - sbir_validation_report.md  — CRITICAL / WARNING / INFO findings
  - eligibility_gates_check.md — gate-level [USER VERIFY] items
  - vol*.md                    — [USER VERIFY] / [PLACEHOLDER] markers
  - per_proposal_lookup.md     — date fields + [USER VERIFY] markers
  - .sbir_validation_<verdict> — current validator marker (verdict + timestamp)

Excel output (4 sheets):
  1. Master         — every open item, one row each
  2. By Owner       — filtered views per owner (Rasheed / EZ / Bola / Breion / Cyber SME / Unassigned)
  3. By Topic       — grouped per topic
  4. Summary        — per-topic stats (verdict, counts, deadline)

Source of truth = the .md files. Excel is a derived view. To mark something
"done," edit the .md file (resolve the placeholder, check the box). Next run
will reflect it. No state preservation in v1.

Invoked by:
  - proposal-tracker agent (manual invocation)
  - PostToolUse hook on .sbir_validation_* writes (automatic)
  - CLI directly: `python tools/proposal_tracker.py [--repo-root PATH]`
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class OpenItem:
    topic: str
    volume: str
    item: str
    category: str
    owner: str
    priority: str
    source_file: str
    source_line: int
    notes: str = ""


@dataclass
class TopicSummary:
    topic: str
    component: str
    verdict: str
    critical: int = 0
    warning: int = 0
    info: int = 0
    deadline: str = ""
    open_items: int = 0
    last_validated: str = ""


# ---------------------------------------------------------------------------
# Categorization heuristics
# ---------------------------------------------------------------------------

OWNER_RULES = [
    # (regex pattern, owner)
    (r"\brasheed\b|saleshub|\bprime\b|\bCMMC\b|SPRS|DD\s*Form\s*2345|DD\s*2345|\bJCP\b|UEI|CAGE|SAM\.gov|DSIP\s*Firm", "Rasheed (SalesHub Prime)"),
    (r"dux\s*machina|TRAIGA|prime\s*fleet|USPS|DOT|value\s*builder|PSG\s*framework|6-block|6\s*block|elite\s*5-lever", "EZ (Dux Machina Sub)"),
    (r"cyber\s*SME|OSCP|OSEP|GPEN|MITRE\s*ATT&CK|pen[\-\s]?test", "Cyber SME [unassigned]"),
    (r"\bbola\b|\bbreion\b|dux\s*vitae\s*capital|DVC\s*advisor", "Bola / Breion (Dux Vitae Capital)"),
    (r"subcontract|teaming\s*agreement", "Rasheed + EZ (joint)"),
]

PRIORITY_KEYWORDS = {
    "CRITICAL": [r"\bCRITICAL\b", r"red\s*flag", r"HARD\s*BLOCKER", r"\bblocker\b", r"\bDISQUALIF", r"\bREJECT"],
    "HIGH": [r"\bHIGH\b", r"\bmust\b", r"\brequired\b", r"USER\s*VERIFY.*before\s*submission", r"submission\s*-blocking"],
    "WARNING": [r"\bWARNING\b", r"warning"],
    "INFO": [r"\bINFO\b", r"\binformational\b", r"recommend"],
}

CATEGORY_RULES = [
    (r"eligibility|gate\s*\d|SBC|small\s*business\s*concern|POW|percentage\s*of\s*work", "Eligibility"),
    (r"PI\b|principal\s*investigator|co-investigator|key\s*personnel|CV|resume|SME\b", "Personnel"),
    (r"vol\s*1|cover\s*sheet|abstract|commercialization\s*summary", "Vol 1 (Cover)"),
    (r"vol\s*2|technical\s*volume|SOW|statement\s*of\s*work", "Vol 2 (Technical)"),
    (r"vol\s*3|cost|labor\s*rate|fringe|overhead|G&A|fee", "Vol 3 (Cost)"),
    (r"vol\s*5|supporting\s*doc|letter\s*of\s*support|DD\s*2345|DD\s*Form\s*2345|ITAR|DFARS", "Vol 5 (Supporting)"),
    (r"vol\s*7|foreign\s*affiliation|webform", "Vol 7 (Foreign)"),
    (r"past\s*performance|past\s*perf|prior\s*work", "Past Performance"),
    (r"subcontract|teaming|consultant", "Subcontract"),
    (r"DSIP|deadline|close\s*date|pre-release|Q&A|submission\s*timing", "Schedule"),
    (r"CMMC|NIST\s*800-171|SPRS|cybersecurity\s*posture", "Compliance / CMMC"),
    (r"UEI|CAGE|SAM\.gov|registration", "Registration"),
    (r"phase\s*III|DFARS\s*252.227-7018|IP\s*retention|commercial", "Phase III / IP"),
]


def infer_owner(text: str) -> str:
    t = text.lower()
    for pattern, owner in OWNER_RULES:
        if re.search(pattern, t, re.IGNORECASE):
            return owner
    return "EZ (default)"


def infer_priority(text: str) -> str:
    for prio in ["CRITICAL", "HIGH", "WARNING", "INFO"]:
        for kw in PRIORITY_KEYWORDS[prio]:
            if re.search(kw, text, re.IGNORECASE):
                return prio
    return "MEDIUM"


def infer_category(text: str, source_file: str) -> str:
    t = text.lower()
    src = source_file.lower()
    # Source file hint first
    if "vol1" in src or "cover_sheet" in src:
        return "Vol 1 (Cover)"
    if "vol2" in src or "technical_draft" in src:
        return "Vol 2 (Technical)"
    if "vol3" in src or "cost_backup" in src:
        return "Vol 3 (Cost)"
    if "vol5" in src or "supporting_docs" in src:
        return "Vol 5 (Supporting)"
    if "vol7" in src or "foreign_affiliations" in src:
        return "Vol 7 (Foreign)"
    if "eligibility" in src:
        return "Eligibility"
    if "lookup" in src:
        return "Schedule / Lookup"
    # Content-based
    for pattern, cat in CATEGORY_RULES:
        if re.search(pattern, t, re.IGNORECASE):
            return cat
    return "General"


# ---------------------------------------------------------------------------
# Source scanners
# ---------------------------------------------------------------------------

def scan_partner_checklist(path: Path, topic: str) -> Iterable[OpenItem]:
    """Find unchecked checkboxes `[ ]` and items in HIGH/CRITICAL/RED sections."""
    if not path.exists():
        return []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except Exception as e:
        print(f"  WARN: could not read {path}: {e}", file=sys.stderr)
        return []
    items = []
    for i, line in enumerate(lines, 1):
        m = re.search(r"^\s*[-*]?\s*\[\s\]\s+(.+?)\s*$", line)
        if m:
            text = m.group(1)
            items.append(OpenItem(
                topic=topic,
                volume="Checklist",
                item=text[:240],
                category=infer_category(text, str(path)),
                owner=infer_owner(text),
                priority=infer_priority(text),
                source_file=path.name,
                source_line=i,
                notes="PARTNER_CHECKLIST unchecked item",
            ))
    return items


def scan_validation_report(path: Path, topic: str) -> tuple[list[OpenItem], TopicSummary]:
    """Extract CRITICAL/WARNING/INFO findings + the verdict line."""
    summary = TopicSummary(topic=topic, component=topic.split("26")[0] if "26" in topic else "?", verdict="UNKNOWN")
    items: list[OpenItem] = []
    if not path.exists():
        return items, summary
    try:
        text = path.read_text(encoding="utf-8")
    except Exception as e:
        print(f"  WARN: could not read {path}: {e}", file=sys.stderr)
        return items, summary

    # Verdict
    vm = re.search(r"(?:overall\s*verdict|verdict)\s*[:|]\s*(\*\*)?(PASS|CONDITIONAL_PASS|FAIL)(\*\*)?", text, re.IGNORECASE)
    if vm:
        summary.verdict = vm.group(2).upper()

    # Counts
    cm = re.search(r"critical[_\s]+findings\s*[:|]?\s*(\d+)", text, re.IGNORECASE)
    if cm:
        summary.critical = int(cm.group(1))
    wm = re.search(r"warning[_\s]+findings\s*[:|]?\s*(\d+)", text, re.IGNORECASE)
    if wm:
        summary.warning = int(wm.group(1))
    im = re.search(r"info[_\s]+findings\s*[:|]?\s*(\d+)", text, re.IGNORECASE)
    if im:
        summary.info = int(im.group(1))

    # Findings — capture bullet items under each severity heading
    for severity in ("CRITICAL", "WARNING", "INFO"):
        section_match = re.search(
            rf"###?\s*{severity}\s*findings.*?(?=###?\s*\w|\Z)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if not section_match:
            continue
        section_text = section_match.group(0)
        for bullet in re.finditer(r"^\s*[-*]\s+(.+?)(?:\n|$)", section_text, re.MULTILINE):
            line = bullet.group(1).strip()
            if not line or line.startswith("["):
                continue
            line_no = text[:bullet.start()].count("\n") + 1
            items.append(OpenItem(
                topic=topic,
                volume="Validator",
                item=line[:240],
                category=infer_category(line, str(path)),
                owner=infer_owner(line),
                priority=severity,
                source_file=path.name,
                source_line=line_no,
                notes=f"Validator {severity} finding",
            ))

    return items, summary


PLACEHOLDER_RE = re.compile(r"\[(?:USER\s+VERIFY|PLACEHOLDER)(?:\s*[:|]\s*(.+?))?\]", re.IGNORECASE)


def scan_placeholder_markers(path: Path, topic: str, volume: str) -> Iterable[OpenItem]:
    """Find [USER VERIFY] and [PLACEHOLDER:] markers in any .md file."""
    if not path.exists():
        return []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except Exception as e:
        print(f"  WARN: could not read {path}: {e}", file=sys.stderr)
        return []
    items = []
    seen = set()
    for i, line in enumerate(lines, 1):
        for m in PLACEHOLDER_RE.finditer(line):
            detail = (m.group(1) or "").strip() or line.strip()
            key = (path.name, detail[:120])
            if key in seen:
                continue
            seen.add(key)
            items.append(OpenItem(
                topic=topic,
                volume=volume,
                item=detail[:240],
                category=infer_category(detail, str(path)),
                owner=infer_owner(detail),
                priority=infer_priority(detail),
                source_file=path.name,
                source_line=i,
                notes="Placeholder / USER VERIFY marker",
            ))
    return items


def detect_marker_file(folder: Path) -> tuple[str, str]:
    """Return (verdict, last_validated_iso) from the .sbir_validation_* marker."""
    for variant in ("pass", "conditional", "fail"):
        marker = folder / f".sbir_validation_{variant}"
        if marker.exists():
            try:
                content = marker.read_text(encoding="utf-8").strip()
                ts_match = re.search(r"\d{4}-\d{2}-\d{2}T?[\d:]*", content)
                ts = ts_match.group(0) if ts_match else ""
                return variant.upper(), ts
            except Exception:
                return variant.upper(), ""
    return "UNVALIDATED", ""


VOL_FILENAME_MAP = {
    "vol1_cover_sheet.md": "Vol 1 (Cover)",
    "vol2_technical_draft.md": "Vol 2 (Technical)",
    "vol3_cost_backup.md": "Vol 3 (Cost)",
    "vol5_supporting_docs.md": "Vol 5 (Supporting)",
    "vol7_foreign_affiliations_answers.md": "Vol 7 (Foreign)",
    "eligibility_gates_check.md": "Eligibility",
    "per_proposal_lookup.md": "Lookup",
    "tpoc_outreach_script.md": "TPOC Outreach",
    "TRACEABILITY_MATRIX.md": "Traceability",
    "rre_structure.md": "RRE",
    "3layer_mapping.md": "3-Layer Map",
    "shred_matrix.md": "Shred",
}


def scan_topic_folder(folder: Path) -> tuple[list[OpenItem], TopicSummary]:
    topic = folder.name
    items: list[OpenItem] = []

    # Validation report → summary + findings
    val_path = folder / "sbir_validation_report.md"
    val_items, summary = scan_validation_report(val_path, topic)
    items.extend(val_items)

    # Marker file overrides verdict if present
    marker_verdict, marker_ts = detect_marker_file(folder)
    if marker_verdict != "UNVALIDATED":
        summary.verdict = marker_verdict
        summary.last_validated = marker_ts

    # Partner checklist → unchecked items
    items.extend(scan_partner_checklist(folder / "PARTNER_CHECKLIST.md", topic))

    # Placeholder markers in all vol files
    for filename, volume in VOL_FILENAME_MAP.items():
        items.extend(scan_placeholder_markers(folder / filename, topic, volume))

    # Capture deadline from per_proposal_lookup.md if present
    lookup = folder / "per_proposal_lookup.md"
    if lookup.exists():
        try:
            txt = lookup.read_text(encoding="utf-8")
            dm = re.search(r"SUBMISSION\s*DEADLINE\s*[:|]?\s*(.+?)(?:\n|$)", txt, re.IGNORECASE)
            if dm:
                summary.deadline = dm.group(1).strip()[:60]
        except Exception:
            pass

    summary.open_items = len(items)
    return items, summary


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

PRIORITY_FILL = {
    "CRITICAL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "HIGH":     PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "WARNING":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "MEDIUM":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "LOW":      PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "INFO":     PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
}

PRIORITY_RANK = {"CRITICAL": 0, "HIGH": 1, "WARNING": 2, "MEDIUM": 3, "LOW": 4, "INFO": 5}


# ---------------------------------------------------------------------------
# Task deduplication
# ---------------------------------------------------------------------------
# Each TASK_DEFINITION groups multiple raw placeholder rows into ONE actionable
# task. Pattern is matched against the item text + notes. First-match wins
# (rows matched by an earlier task are excluded from later tasks).
# Tune these patterns as new proposals reveal new task themes.

TASK_DEFINITIONS = [
    # (display name, owner, priority, regex pattern)
    ("Execute SalesHub <-> Dux Machina subcontract / teaming agreement",
     "Rasheed + EZ (joint)", "CRITICAL",
     r"subcontract|teaming\s*agreement"),

    ("Confirm SalesHub DD Form 2345 active in JCP",
     "Rasheed (SalesHub Prime)", "CRITICAL",
     r"DD\s*Form\s*2345|DD\s*2345|\bJCP\b"),

    ("Provide Rasheed Chowdhury's CV / PI bio",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"rasheed.*(cv|resume|bio)|chowdhury.*(cv|resume|bio)|PI\s*(bio|CV|resume)"),

    ("Provide SalesHub UEI + CAGE codes",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"\bUEI\b|\bCAGE\b"),

    ("Provide 2-3 SalesHub federal past-performance references",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"saleshub.*past.*perf|salesforce.*past.*perf|TracAnything|NPS.*Survey|saleshub.*federal"),

    ("Verify SalesHub CMMC L2 SPRS posting current",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"CMMC.*SPRS|SPRS.*post|SPRS.*current|800-?171"),

    ("Rasheed-ratified Vol 7 Foreign Affiliations answers (all 8 Qs)",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"foreign\s*affiliation|18\s*USC\s*1001|vol\s*7"),

    ("Finalize Vol 3 labor rates (loaded rates + fringe/OH/G&A/fee)",
     "Rasheed (SalesHub Prime)", "HIGH",
     r"labor\s*rate|fringe|overhead|G&A|fee\s*rate|loaded\s*rate|hourly\s*rate|indirect\s*rate"),

    ("Capture DSIP Release 2 dates (submission / Q&A / pre-release)",
     "EZ (default)", "HIGH",
     r"DSIP.*deadline|DSIP.*close|pre-release|Q&A.*close|submission.*deadline"),

    ("Pursue DLA J6 CIO Letter of Support (NV004 only)",
     "EZ (default)", "HIGH",
     r"J6.*CIO|letter\s*of\s*support|\bLOS\b"),

    ("Identify and onboard Cyber SME (OSCP/GPEN credentials) - NV005",
     "Rasheed + EZ (joint)", "HIGH",
     r"cyber\s*SME|\bOSCP\b|\bGPEN\b|\bOSEP\b|MITRE\s*ATT&CK"),

    ("Quantify Dux Machina past-perf metrics ($/dates/contract refs)",
     "EZ (Dux Machina Sub)", "MEDIUM",
     r"prime\s*fleet|TRAIGA|\bUSPS\b|\bDOT\b|value\s*builder|contract\s*number|contract\s*vehicle"),
]


def dedupe_into_tasks(items: list[OpenItem]) -> list[dict]:
    """Group raw placeholder rows into actionable tasks via TASK_DEFINITIONS.
    Unmatched rows are bucketed into 'Other [category]' rows per topic.
    Returns a list of task dicts ready for the Tasks sheet."""
    tasks = []
    matched_idx = set()

    for name, owner, priority, pattern in TASK_DEFINITIONS:
        pat = re.compile(pattern, re.IGNORECASE)
        matches = []
        for idx, it in enumerate(items):
            if idx in matched_idx:
                continue
            haystack = f"{it.item} {it.notes}".lower()
            if pat.search(haystack):
                matches.append(it)
                matched_idx.add(idx)
        if not matches:
            continue
        topics = sorted(set(m.topic for m in matches))
        files = sorted(set(m.source_file for m in matches))
        tasks.append({
            "name": name,
            "owner": owner,
            "topics": ", ".join(topics),
            "priority": priority,
            "placeholders": len(matches),
            "files": ", ".join(files),
            "status": "OPEN",
        })

    # Catch-all for unmatched placeholders, grouped by (topic, category)
    from collections import defaultdict
    unmatched = [items[i] for i in range(len(items)) if i not in matched_idx]
    groups = defaultdict(list)
    for it in unmatched:
        groups[(it.topic, it.category)].append(it)
    for (topic, category), group in sorted(groups.items()):
        if len(group) < 1:
            continue
        tasks.append({
            "name": f"Other [{category}] items - {topic}",
            "owner": "Mixed (review individually)",
            "topics": topic,
            "priority": "MEDIUM",
            "placeholders": len(group),
            "files": ", ".join(sorted(set(g.source_file for g in group))),
            "status": "OPEN",
        })

    tasks.sort(key=lambda t: (PRIORITY_RANK.get(t["priority"], 9), t["name"]))
    return tasks


def autosize(ws, max_width=80):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_excel(
    items: list[OpenItem],
    summaries: list[TopicSummary],
    out_path: Path,
) -> None:
    wb = Workbook()

    # ---- Sheet 1: Tasks (deduplicated to-do list — opens by default) ----
    ws_tasks = wb.active
    ws_tasks.title = "Tasks"
    ws_tasks.append(["Task", "Owner", "Topics Affected", "Priority", "# Placeholders", "Files", "Status"])
    for c in ws_tasks[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")
    tasks = dedupe_into_tasks(items)
    for t in tasks:
        ws_tasks.append([
            t["name"], t["owner"], t["topics"], t["priority"],
            t["placeholders"], t["files"], t["status"]
        ])
        if t["priority"] in PRIORITY_FILL:
            ws_tasks.cell(row=ws_tasks.max_row, column=4).fill = PRIORITY_FILL[t["priority"]]
    ws_tasks.freeze_panes = "A2"
    ws_tasks.auto_filter.ref = ws_tasks.dimensions
    autosize(ws_tasks, max_width=80)

    # ---- Sheet 2: Master ----
    ws = wb.create_sheet("Master")
    headers = ["Topic", "Volume", "Item", "Category", "Owner", "Priority", "Source File", "Line", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")
    sorted_items = sorted(items, key=lambda x: (PRIORITY_RANK.get(x.priority, 9), x.topic, x.category))
    for it in sorted_items:
        ws.append([it.topic, it.volume, it.item, it.category, it.owner, it.priority, it.source_file, it.source_line, it.notes])
        if it.priority in PRIORITY_FILL:
            ws.cell(row=ws.max_row, column=6).fill = PRIORITY_FILL[it.priority]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)

    # ---- Sheet 2: By Owner ----
    ws2 = wb.create_sheet("By Owner")
    ws2.append(["Owner", "Topic", "Priority", "Item", "Category", "Source File"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for it in sorted(sorted_items, key=lambda x: (x.owner, PRIORITY_RANK.get(x.priority, 9))):
        ws2.append([it.owner, it.topic, it.priority, it.item, it.category, it.source_file])
        if it.priority in PRIORITY_FILL:
            ws2.cell(row=ws2.max_row, column=3).fill = PRIORITY_FILL[it.priority]
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    autosize(ws2)

    # ---- Sheet 3: By Topic ----
    ws3 = wb.create_sheet("By Topic")
    ws3.append(["Topic", "Priority", "Owner", "Item", "Category", "Volume", "Source File"])
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for it in sorted(sorted_items, key=lambda x: (x.topic, PRIORITY_RANK.get(x.priority, 9))):
        ws3.append([it.topic, it.priority, it.owner, it.item, it.category, it.volume, it.source_file])
        if it.priority in PRIORITY_FILL:
            ws3.cell(row=ws3.max_row, column=2).fill = PRIORITY_FILL[it.priority]
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    autosize(ws3)

    # ---- Sheet 4: Summary ----
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Topic", "Component", "Verdict", "Critical", "Warning", "Info", "Total Open Items", "Deadline", "Last Validated"])
    for c in ws4[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for s in sorted(summaries, key=lambda x: x.topic):
        ws4.append([s.topic, s.component, s.verdict, s.critical, s.warning, s.info, s.open_items, s.deadline, s.last_validated])
        verdict_fill = {
            "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "CONDITIONAL_PASS": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }.get(s.verdict)
        if verdict_fill:
            ws4.cell(row=ws4.max_row, column=3).fill = verdict_fill
    ws4.freeze_panes = "A2"
    autosize(ws4)

    # ---- Metadata row at bottom of Summary ----
    ws4.append([])
    ws4.append([f"Generated: {datetime.now().isoformat(timespec='seconds')}"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(italic=True, color="808080")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build/refresh PROPOSAL_TRACKER.xlsx")
    parser.add_argument(
        "--repo-root",
        default=r"C:\Users\sabaa\ONEDRIVE\DESKTOP\TEST_AGENTS",
        help="Repository root (default: TEST_AGENTS)",
    )
    parser.add_argument("--quiet", action="store_true", help="Suppress per-folder output")
    args = parser.parse_args()

    repo_root = Path(args.repo_root)
    outputs_dir = repo_root / "PROPOSAL_TEAM" / "outputs"
    excel_path = outputs_dir / "PROPOSAL_TRACKER.xlsx"

    if not outputs_dir.exists():
        print(f"ERROR: outputs directory not found: {outputs_dir}", file=sys.stderr)
        sys.exit(1)

    # Topic folders = subfolders matching SBIR topic-ID-like pattern
    topic_pattern = re.compile(r"^[A-Z]{3,6}\d{2}B[XZ]\d{2}-[ND]V\d{3}$", re.IGNORECASE)
    topic_folders = sorted([
        p for p in outputs_dir.iterdir()
        if p.is_dir() and topic_pattern.match(p.name)
    ])

    if not topic_folders:
        print(f"WARN: no topic folders found in {outputs_dir}", file=sys.stderr)

    all_items: list[OpenItem] = []
    all_summaries: list[TopicSummary] = []

    for folder in topic_folders:
        if not args.quiet:
            print(f"Scanning {folder.name}/")
        items, summary = scan_topic_folder(folder)
        all_items.extend(items)
        all_summaries.append(summary)
        if not args.quiet:
            print(f"  {len(items)} open items, verdict={summary.verdict}")

    write_excel(all_items, all_summaries, excel_path)
    print(f"OK: wrote {excel_path}")
    print(f"     {len(all_items)} open items across {len(all_summaries)} proposal(s)")


if __name__ == "__main__":
    main()
