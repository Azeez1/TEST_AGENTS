"""
TRIBE v2 Neural Content Optimization Calibrator
===============================================
Runs videos through TRIBE v2 (Meta FAIR) + VidIQ + YouTube Most Replayed,
computes calibration metrics, persists to JSON + xlsx.

Usage:
    python tribev2_calibrator.py --url <YOUTUBE_URL>
    python tribev2_calibrator.py --batch urls.txt
    python tribev2_calibrator.py --decision-card
    python tribev2_calibrator.py --reset-dataset

License note:
    TRIBE v2 is CC BY-NC 4.0. This tool is for internal Dux Machina
    research use only. Do NOT use for commercial client work without
    separate licensing from Meta.
"""

from __future__ import annotations

import argparse
import json
import logging
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

# ── Path constants ───────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
TEAM_ROOT = SCRIPT_DIR.parent
NEURAL_CAL_DIR = TEAM_ROOT / "outputs" / "neural-calibration"
DATASET_JSON = NEURAL_CAL_DIR / "calibration-dataset.json"
DATASET_XLSX = NEURAL_CAL_DIR / "calibration-dataset.xlsx"
VIDEOS_DIR = NEURAL_CAL_DIR / "videos"
TRIBE_DIR = NEURAL_CAL_DIR / "tribe-outputs"
LOGS_DIR = NEURAL_CAL_DIR / "logs"

# ── Go/No-Go thresholds (LOCKED before data collection) ──────────────────

GO_PEAK_OVERLAP = 0.70
GO_PEARSON_R = 0.50
GO_MIN_VIDEOS = 7
REFINE_PEAK_OVERLAP_MIN = 0.50
REFINE_PEARSON_R_MIN = 0.30


def setup_logger() -> logging.Logger:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOGS_DIR / f"calibrator-{datetime.now():%Y-%m-%d}.log"
    logger = logging.getLogger("tribev2_calibrator")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(fmt)
    logger.addHandler(console)

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)

    return logger


log = setup_logger()


# ── Stage functions (stubs, filled in later tasks) ───────────────────────

def download_video(url: str) -> dict:
    """Download video with yt-dlp. Returns metadata dict."""
    log.info("downloading: %s", url)

    output_template = str(VIDEOS_DIR / "%(id)s.%(ext)s")
    cmd = [
        "yt-dlp",
        "-o", output_template,
        "--write-info-json",
        "--no-playlist",
        "--max-filesize", "500M",
        "--quiet",
        "--no-warnings",
        url,
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
    except subprocess.TimeoutExpired:
        log.error("yt-dlp timeout after 10 min")
        return {"error": "download_timeout", "url": url}

    if result.returncode != 0:
        log.error("yt-dlp failed: %s", result.stderr.strip()[:500])
        return {"error": "download_failed", "message": result.stderr.strip()[:500], "url": url}

    info_files = list(VIDEOS_DIR.glob("*.info.json"))
    matching = [f for f in info_files if url.endswith(f.stem.replace(".info", ""))]
    if not matching:
        info_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        matching = info_files[:1]

    if not matching:
        return {"error": "info_json_not_found", "url": url}

    info_path = matching[0]
    with info_path.open(encoding="utf-8") as f:
        info = json.load(f)

    video_id = info.get("id", "")
    video_files = list(VIDEOS_DIR.glob(f"{video_id}.*"))
    video_files = [f for f in video_files if not f.name.endswith(".info.json")]
    local_path = str(video_files[0]) if video_files else ""

    return {
        "video_id": video_id,
        "title": info.get("title", ""),
        "duration": info.get("duration", 0),
        "view_count": info.get("view_count", 0),
        "upload_date": info.get("upload_date", ""),
        "channel": info.get("uploader") or info.get("channel", ""),
        "url": url,
        "local_path": local_path,
    }


def run_tribe_v2(video_path: Path, video_id: str) -> dict:
    """
    Trigger TRIBE v2 analysis via Colab with manual bridge.

    Workflow:
      1. Tool prints instructions to upload video to Colab
      2. User runs TRIBE v2 in Colab, exports result JSON
      3. User saves result to tribe-outputs/{video_id}.json
      4. User presses ENTER
      5. Tool reads the JSON and returns parsed dict
    """
    expected_path = TRIBE_DIR / f"{video_id}.json"

    print()
    print("=" * 72)
    print(f"TRIBE v2 MANUAL STEP for video: {video_id}")
    print("=" * 72)
    print(f"  1. Upload this file to your Colab notebook:")
    print(f"     {video_path}")
    print(f"  2. Run the TRIBE v2 cells in Colab on this video")
    print(f"  3. Export the result JSON with these keys:")
    print(f"       activation_timeline  (list of floats, 1 per second)")
    print(f"       peak_moments         (list of timestamps in seconds)")
    print(f"       dead_zones           (list of timestamps in seconds)")
    print(f"       aggregate_score      (float, 0.0 to 1.0)")
    print(f"       multimodal_score     (float, optional)")
    print(f"  4. Save the result JSON to:")
    print(f"     {expected_path}")
    print()
    user_input = input("Press ENTER when ready (or type 'skip' to skip this video): ").strip().lower()

    if user_input == "skip":
        log.warning("user skipped tribe v2 for %s", video_id)
        return {"skipped": True, "reason": "user_skipped"}

    if not expected_path.exists():
        log.error("tribe output file not found: %s", expected_path)
        return {"error": "file_not_found", "expected_path": str(expected_path)}

    try:
        with expected_path.open(encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        log.error("tribe output parse failed: %s", e)
        return {"error": "parse_failed", "message": str(e)}

    required_keys = ["activation_timeline", "peak_moments", "dead_zones", "aggregate_score"]
    missing = [k for k in required_keys if k not in data]
    if missing:
        log.warning("tribe output missing keys: %s", missing)

    return {
        "activation_timeline": data.get("activation_timeline", []),
        "peak_moments": data.get("peak_moments", []),
        "dead_zones": data.get("dead_zones", []),
        "aggregate_score": float(data.get("aggregate_score", 0.0)),
        "multimodal_score": float(data.get("multimodal_score", 0.0)) if "multimodal_score" in data else None,
        "modalities_used": data.get("modalities_used", ["video", "audio", "text"]),
        "source_file": str(expected_path),
    }


def _safe_float(prompt: str, default: float | None = None) -> float | None:
    """Prompt for a numeric value, return default on skip/empty/invalid."""
    raw = input(prompt).strip()
    if not raw or raw.lower() == "skip":
        return default
    try:
        return float(raw)
    except ValueError:
        log.warning("invalid number '%s', using default %s", raw, default)
        return default


def fetch_vidiq_metrics(url: str, video_id: str) -> dict:
    """Capture VidIQ public metrics via manual prompt."""
    print()
    print("-" * 72)
    print(f"VidIQ METRICS for video: {video_id}")
    print("-" * 72)
    print(f"  Open in browser: {url}")
    print(f"  Let the VidIQ extension populate, then paste values below.")
    print(f"  Press ENTER (or type 'skip') to skip a field.")
    print()

    engagement_rate = _safe_float("  Engagement rate (decimal, e.g. 0.045 for 4.5%): ")
    views_per_hour = _safe_float("  Views per hour: ")
    performance_score = _safe_float("  VidIQ performance score (0-100): ")

    captured = {
        "engagement_rate": engagement_rate,
        "views_per_hour": views_per_hour,
        "performance_score": performance_score,
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "capture_mode": "manual",
    }

    all_skipped = all(v is None for v in (engagement_rate, views_per_hour, performance_score))
    if all_skipped:
        return {"skipped": True, "reason": "all_fields_empty"}

    return captured


def fetch_most_replayed(url: str, video_id: str) -> dict:
    """
    Scrape YouTube Most Replayed heatmap from the watch page.
    Returns {'available': bool, 'heatmap': [...], 'peaks': [...]}
    """
    import re
    import requests

    try:
        response = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            },
            timeout=30,
        )
        response.raise_for_status()
    except Exception as e:
        log.warning("most replayed fetch failed: %s", e)
        return {"available": False, "error": str(e)[:200]}

    html = response.text

    heatmap_match = re.search(
        r'"markerType":"MARKER_TYPE_HEATMAP","markers":(\[.*?\])',
        html,
        re.DOTALL,
    )

    if not heatmap_match:
        return {"available": False, "reason": "no_heatmap_in_page"}

    try:
        markers = json.loads(heatmap_match.group(1))
    except json.JSONDecodeError as e:
        return {"available": False, "reason": f"heatmap_parse_failed: {e}"}

    heatmap = []
    for marker in markers:
        start_ms = int(marker.get("startMillis", 0))
        intensity = float(marker.get("intensityScoreNormalized", 0.0))
        heatmap.append([start_ms / 1000.0, intensity])

    if not heatmap:
        return {"available": False, "reason": "empty_heatmap"}

    sorted_by_intensity = sorted(heatmap, key=lambda x: x[1], reverse=True)
    top_n = max(3, len(heatmap) // 5)
    peaks = sorted([p[0] for p in sorted_by_intensity[:top_n]])

    return {
        "available": True,
        "heatmap": heatmap,
        "peaks": peaks,
        "marker_count": len(heatmap),
    }


def _peak_overlap_pct(tribe_peaks: list, replay_peaks: list, tolerance_s: float = 3.0) -> float | None:
    """Percentage of TRIBE peaks landing within tolerance of any Most Replayed peak."""
    if not tribe_peaks:
        return None
    if not replay_peaks:
        return None
    matched = sum(
        1 for tp in tribe_peaks
        if any(abs(float(tp) - float(rp)) <= tolerance_s for rp in replay_peaks)
    )
    return matched / len(tribe_peaks)


def _pearson_r(tribe_timeline: list, heatmap: list, video_duration: float) -> float | None:
    """
    Pearson correlation between TRIBE activation timeline and Most Replayed heatmap.
    Both are interpolated to a common per-second grid first.
    """
    import numpy as np
    from scipy.stats import pearsonr

    if not tribe_timeline or not heatmap:
        return None

    duration = max(int(video_duration), len(tribe_timeline))
    if duration <= 2:
        return None

    grid_seconds = np.arange(duration, dtype=float)

    tribe_arr = np.array(tribe_timeline, dtype=float)
    if len(tribe_arr) != duration:
        xs_tribe = np.linspace(0, duration - 1, len(tribe_arr))
        tribe_interp = np.interp(grid_seconds, xs_tribe, tribe_arr)
    else:
        tribe_interp = tribe_arr

    replay_times = np.array([p[0] for p in heatmap], dtype=float)
    replay_values = np.array([p[1] for p in heatmap], dtype=float)
    if len(replay_times) < 2:
        return None
    replay_interp = np.interp(grid_seconds, replay_times, replay_values)

    if np.std(tribe_interp) < 1e-9 or np.std(replay_interp) < 1e-9:
        return None

    try:
        r, _ = pearsonr(tribe_interp, replay_interp)
        return float(r)
    except Exception as e:
        log.warning("pearson failed: %s", e)
        return None


def compute_calibration_metrics(record: dict) -> dict:
    """
    Compute peak overlap percentage, Pearson correlation, and composite accuracy.
    Returns a dict with decision flag per video.
    """
    tribe = record.get("tribe_v2") or {}
    most_replayed = record.get("most_replayed") or {}

    tribe_peaks = tribe.get("peak_moments", [])
    tribe_timeline = tribe.get("activation_timeline", [])
    replay_peaks = most_replayed.get("peaks", []) if most_replayed.get("available") else []
    heatmap = most_replayed.get("heatmap", []) if most_replayed.get("available") else []
    duration = record.get("duration_seconds", 0)

    peak_overlap = _peak_overlap_pct(tribe_peaks, replay_peaks)
    pearson = _pearson_r(tribe_timeline, heatmap, duration)

    composite = 0.0
    if peak_overlap is not None:
        composite += 0.6 * peak_overlap
    if pearson is not None:
        composite += 0.4 * max(pearson, 0.0)
    if peak_overlap is None and pearson is None:
        composite = None

    if composite is None:
        decision_flag = "insufficient_data"
    elif composite >= 0.70:
        decision_flag = "go"
    elif composite >= 0.40:
        decision_flag = "refine"
    else:
        decision_flag = "pivot"

    return {
        "peak_overlap_pct": peak_overlap,
        "pearson_r": pearson,
        "composite_accuracy_score": composite,
        "decision_flag": decision_flag,
        "method": "0.6 * peak_overlap + 0.4 * max(pearson_r, 0)",
    }


def create_empty_record(video_meta: dict, category: str) -> dict:
    """Build a fresh calibration record with null fields for downstream stages."""
    now = datetime.now(timezone.utc).isoformat()
    return {
        "run_id": f"run_{now}_{video_meta.get('video_id', 'unknown')}",
        "video_id": video_meta.get("video_id", ""),
        "url": video_meta.get("url", ""),
        "title": video_meta.get("title", ""),
        "channel": video_meta.get("channel", ""),
        "upload_date": video_meta.get("upload_date", ""),
        "duration_seconds": video_meta.get("duration", 0),
        "view_count": video_meta.get("view_count", 0),
        "local_path": video_meta.get("local_path", ""),
        "category": category,
        "analyzed_at": now,
        "tribe_v2": None,
        "vidiq": None,
        "most_replayed": None,
        "calibration": None,
        "errors": [],
        "notes": "",
    }


def load_dataset() -> list:
    """Load the calibration dataset JSON. Returns [] if missing or empty."""
    if not DATASET_JSON.exists():
        return []
    try:
        with DATASET_JSON.open(encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        log.warning("dataset JSON corrupted, starting fresh")
        return []


def save_dataset(dataset: list) -> None:
    """Write the dataset atomically with a rolling backup."""
    if DATASET_JSON.exists():
        backup = DATASET_JSON.with_suffix(".json.bak")
        backup.write_bytes(DATASET_JSON.read_bytes())

    tmp = DATASET_JSON.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(dataset, f, indent=2, ensure_ascii=False)
    tmp.replace(DATASET_JSON)


def append_to_dataset(record: dict) -> None:
    """Append record to JSON dataset with rolling backup."""
    dataset = load_dataset()
    existing_ids = {r.get("run_id") for r in dataset}
    if record.get("run_id") in existing_ids:
        log.warning("duplicate run_id, skipping: %s", record.get("run_id"))
        return
    dataset.append(record)
    save_dataset(dataset)
    log.info("appended record %s (total: %d)", record.get("video_id"), len(dataset))
    try:
        regenerate_xlsx_from_json()
    except Exception as e:
        log.warning("xlsx regeneration failed: %s", e)


XLSX_COLUMNS = [
    "run_id",
    "video_id",
    "title",
    "channel",
    "category",
    "duration_s",
    "view_count",
    "tribe_aggregate_score",
    "tribe_peaks_count",
    "vidiq_engagement_rate",
    "vidiq_views_per_hour",
    "vidiq_performance_score",
    "most_replayed_available",
    "peak_overlap_pct",
    "pearson_r",
    "composite_score",
    "decision_flag",
    "analyzed_at",
    "errors",
]


def _flatten_record(record: dict) -> list:
    """Convert a nested record into a flat row matching XLSX_COLUMNS."""
    tribe = record.get("tribe_v2") or {}
    vidiq = record.get("vidiq") or {}
    most_replayed = record.get("most_replayed") or {}
    calibration = record.get("calibration") or {}
    errors = record.get("errors") or []

    return [
        record.get("run_id", ""),
        record.get("video_id", ""),
        record.get("title", ""),
        record.get("channel", ""),
        record.get("category", ""),
        record.get("duration_seconds", ""),
        record.get("view_count", ""),
        tribe.get("aggregate_score", ""),
        len(tribe.get("peak_moments", [])) if tribe else "",
        vidiq.get("engagement_rate", ""),
        vidiq.get("views_per_hour", ""),
        vidiq.get("performance_score", ""),
        most_replayed.get("available", ""),
        calibration.get("peak_overlap_pct", ""),
        calibration.get("pearson_r", ""),
        calibration.get("composite_accuracy_score", ""),
        calibration.get("decision_flag", ""),
        record.get("analyzed_at", ""),
        "; ".join(str(e) for e in errors) if errors else "",
    ]


def regenerate_xlsx_from_json() -> None:
    """Rebuild the xlsx spreadsheet from the JSON source of truth."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    dataset = load_dataset()

    wb = Workbook()
    ws = wb.active
    ws.title = "Calibration"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, name in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, record in enumerate(dataset, start=2):
        for col_idx, value in enumerate(_flatten_record(record), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    if dataset:
        summary_row = len(dataset) + 3
        ws.cell(row=summary_row, column=1, value="AVERAGES").font = Font(bold=True)
        numeric_cols = {
            "tribe_aggregate_score": 8,
            "vidiq_engagement_rate": 10,
            "vidiq_views_per_hour": 11,
            "vidiq_performance_score": 12,
            "peak_overlap_pct": 14,
            "pearson_r": 15,
            "composite_score": 16,
        }
        for name, col_idx in numeric_cols.items():
            col_letter = get_column_letter(col_idx)
            formula = f"=IFERROR(AVERAGE({col_letter}2:{col_letter}{len(dataset)+1}),\"\")"
            ws.cell(row=summary_row, column=col_idx, value=formula).font = Font(bold=True)

    for col_idx, name in enumerate(XLSX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(name) + 2)

    ws.freeze_panes = "A2"

    wb.save(DATASET_XLSX)
    log.info("regenerated xlsx (%d rows)", len(dataset))


def decision_card() -> dict:
    """
    Compute aggregate metrics across the full dataset and apply GO/REFINE/PIVOT thresholds.
    Writes a markdown report and returns the decision dict.
    """
    import numpy as np

    dataset = load_dataset()
    if not dataset:
        log.warning("dataset empty, cannot compute decision")
        return {"verdict": "no_data", "message": "Run some videos first."}

    valid_records = []
    for rec in dataset:
        calibration = rec.get("calibration") or {}
        composite = calibration.get("composite_accuracy_score")
        if composite is not None:
            valid_records.append(rec)

    if not valid_records:
        log.warning("no valid records with calibration data")
        return {"verdict": "no_data", "message": "No videos have complete calibration data yet."}

    peak_overlaps = [
        (r["calibration"] or {}).get("peak_overlap_pct")
        for r in valid_records
        if (r["calibration"] or {}).get("peak_overlap_pct") is not None
    ]
    pearsons = [
        (r["calibration"] or {}).get("pearson_r")
        for r in valid_records
        if (r["calibration"] or {}).get("pearson_r") is not None
    ]
    composites = [
        (r["calibration"] or {}).get("composite_accuracy_score")
        for r in valid_records
    ]

    avg_peak_overlap = float(np.mean(peak_overlaps)) if peak_overlaps else None
    avg_pearson = float(np.mean(pearsons)) if pearsons else None
    avg_composite = float(np.mean(composites))

    go_count = sum(
        1 for r in valid_records
        if (r["calibration"] or {}).get("decision_flag") == "go"
    )
    refine_count = sum(
        1 for r in valid_records
        if (r["calibration"] or {}).get("decision_flag") == "refine"
    )
    pivot_count = sum(
        1 for r in valid_records
        if (r["calibration"] or {}).get("decision_flag") == "pivot"
    )

    n = len(valid_records)

    def in_range(value, low, high):
        return value is not None and low <= value < high

    passes_go = (
        avg_peak_overlap is not None
        and avg_pearson is not None
        and avg_peak_overlap >= GO_PEAK_OVERLAP
        and avg_pearson >= GO_PEARSON_R
        and go_count >= GO_MIN_VIDEOS
    )
    passes_refine = (
        (avg_peak_overlap is not None and in_range(avg_peak_overlap, REFINE_PEAK_OVERLAP_MIN, GO_PEAK_OVERLAP))
        or (avg_pearson is not None and in_range(avg_pearson, REFINE_PEARSON_R_MIN, GO_PEARSON_R))
    )

    if passes_go:
        verdict = "GO"
        next_actions = [
            "Start offering Neural Content Optimization as a paid Dux Machina service",
            "Create a pitch deck highlighting the calibration results",
            "Reach out to the Paperclip/Iron Grove network for pilot clients",
            "Package the workflow into a client onboarding document",
        ]
    elif passes_refine:
        verdict = "REFINE"
        next_actions = [
            "Run another batch of 10 videos with tighter methodology",
            "Review which videos scored lowest to identify weak signal sources",
            "Consider using longer-form videos (TRIBE may perform better on >60s content)",
            "Test with videos from different niches to control for topic bias",
        ]
    else:
        verdict = "PIVOT"
        next_actions = [
            "TRIBE v2 predictions don't correlate strongly with public engagement signals",
            "Repackage TRIBE v2 as an internal research tool only, not a client service",
            "Consider alternate uses: pre-launch video brainstorming, creative direction input",
            "Focus on the autoresearch + audit diagnosis playbook which already has proven ROI",
        ]

    card = {
        "verdict": verdict,
        "videos_analyzed": n,
        "avg_peak_overlap_pct": avg_peak_overlap,
        "avg_pearson_r": avg_pearson,
        "avg_composite_score": avg_composite,
        "videos_by_flag": {
            "go": go_count,
            "refine": refine_count,
            "pivot": pivot_count,
        },
        "thresholds": {
            "GO": f"avg_peak_overlap >= {GO_PEAK_OVERLAP} AND avg_pearson_r >= {GO_PEARSON_R} AND {GO_MIN_VIDEOS}+ videos flagged go",
            "REFINE": f"avg_peak_overlap in [{REFINE_PEAK_OVERLAP_MIN}, {GO_PEAK_OVERLAP}) OR avg_pearson_r in [{REFINE_PEARSON_R_MIN}, {GO_PEARSON_R})",
            "PIVOT": "anything below REFINE",
        },
        "next_actions": next_actions,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    _print_decision_card(card)
    _save_decision_card_md(card)

    return card


def _print_decision_card(card: dict) -> None:
    """Pretty-print the decision card to console."""
    def fmt(v):
        if v is None:
            return "N/A"
        if isinstance(v, float):
            return f"{v:.3f}"
        return str(v)

    print()
    print("=" * 72)
    print("NEURAL CALIBRATION DECISION CARD")
    print("=" * 72)
    print(f"Videos analyzed:       {card['videos_analyzed']}")
    print(f"Avg peak overlap:      {fmt(card['avg_peak_overlap_pct'])}")
    print(f"Avg Pearson r:         {fmt(card['avg_pearson_r'])}")
    print(f"Avg composite score:   {fmt(card['avg_composite_score'])}")
    print()
    print("Flag distribution:")
    for flag, count in card["videos_by_flag"].items():
        print(f"  {flag.upper():8s} {count}")
    print()
    print(f"VERDICT: {card['verdict']}")
    print()
    print("Next actions:")
    for i, action in enumerate(card["next_actions"], 1):
        print(f"  {i}. {action}")
    print("=" * 72)


def _save_decision_card_md(card: dict) -> None:
    """Save the decision card as markdown for future reference."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    md_path = NEURAL_CAL_DIR / f"decision-card-{date_str}.md"

    def fmt(v):
        if v is None:
            return "N/A"
        if isinstance(v, float):
            return f"{v:.3f}"
        return str(v)

    lines = [
        f"# Neural Calibration Decision Card",
        f"",
        f"**Generated:** {card['generated_at']}",
        f"**Videos analyzed:** {card['videos_analyzed']}",
        f"",
        f"## Aggregate Metrics",
        f"",
        f"| Metric | Value |",
        f"|---|---|",
        f"| Avg peak overlap | {fmt(card['avg_peak_overlap_pct'])} |",
        f"| Avg Pearson r | {fmt(card['avg_pearson_r'])} |",
        f"| Avg composite score | {fmt(card['avg_composite_score'])} |",
        f"",
        f"## Flag Distribution",
        f"",
        f"| Flag | Count |",
        f"|---|---|",
        f"| GO | {card['videos_by_flag']['go']} |",
        f"| REFINE | {card['videos_by_flag']['refine']} |",
        f"| PIVOT | {card['videos_by_flag']['pivot']} |",
        f"",
        f"## Thresholds",
        f"",
        f"- **GO:** {card['thresholds']['GO']}",
        f"- **REFINE:** {card['thresholds']['REFINE']}",
        f"- **PIVOT:** {card['thresholds']['PIVOT']}",
        f"",
        f"## Verdict: **{card['verdict']}**",
        f"",
        f"### Next Actions",
        f"",
    ]
    for action in card["next_actions"]:
        lines.append(f"- {action}")
    lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    log.info("decision card saved: %s", md_path)


# ── Pipeline orchestration ───────────────────────────────────────────────

def process_single_url(url: str, category: str = "uncategorized") -> dict:
    """Run the full TRIBE v2 calibration pipeline on one video URL."""
    log.info("=" * 72)
    log.info("pipeline start: %s (%s)", url, category)
    log.info("=" * 72)

    meta = download_video(url)
    if "error" in meta:
        log.error("download failed: %s", meta.get("error"))
        record = create_empty_record({"url": url}, category)
        record["errors"].append(f"download: {meta.get('error')}")
        append_to_dataset(record)
        return record

    record = create_empty_record(meta, category)

    try:
        tribe_result = run_tribe_v2(Path(meta["local_path"]), meta["video_id"])
        if "error" in tribe_result:
            record["errors"].append(f"tribe_v2: {tribe_result['error']}")
        elif tribe_result.get("skipped"):
            record["errors"].append("tribe_v2: skipped by user")
        else:
            record["tribe_v2"] = tribe_result
    except Exception as e:
        log.exception("tribe_v2 stage failed")
        record["errors"].append(f"tribe_v2: {type(e).__name__}: {e}")

    try:
        vidiq_result = fetch_vidiq_metrics(url, meta["video_id"])
        if vidiq_result.get("skipped"):
            record["errors"].append("vidiq: skipped by user")
        else:
            record["vidiq"] = vidiq_result
    except Exception as e:
        log.exception("vidiq stage failed")
        record["errors"].append(f"vidiq: {type(e).__name__}: {e}")

    try:
        most_replayed_result = fetch_most_replayed(url, meta["video_id"])
        record["most_replayed"] = most_replayed_result
    except Exception as e:
        log.exception("most_replayed stage failed")
        record["errors"].append(f"most_replayed: {type(e).__name__}: {e}")
        record["most_replayed"] = {"available": False, "error": str(e)[:200]}

    try:
        calibration = compute_calibration_metrics(record)
        record["calibration"] = calibration
        log.info(
            "calibration: peak_overlap=%s pearson_r=%s composite=%s flag=%s",
            calibration.get("peak_overlap_pct"),
            calibration.get("pearson_r"),
            calibration.get("composite_accuracy_score"),
            calibration.get("decision_flag"),
        )
    except Exception as e:
        log.exception("calibration stage failed")
        record["errors"].append(f"calibration: {type(e).__name__}: {e}")

    append_to_dataset(record)
    log.info("pipeline complete: %s", meta["video_id"])
    return record


def process_batch(urls_file: Path) -> None:
    """Run the pipeline on a batch of URLs from a text file."""
    if not urls_file.exists():
        log.error("urls file not found: %s", urls_file)
        return

    entries = []
    for raw_line in urls_file.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            url_part, _, comment = line.partition("#")
            url = url_part.strip()
            category = comment.strip()
        else:
            url = line
            category = "uncategorized"
        if category not in ("viral", "competitor", "anti_example", "uncategorized"):
            category = "uncategorized"
        entries.append((url, category))

    log.info("batch starting: %d URLs", len(entries))

    for idx, (url, category) in enumerate(entries, start=1):
        log.info("[%d/%d] processing %s", idx, len(entries), url)
        try:
            process_single_url(url, category)
        except KeyboardInterrupt:
            log.warning("batch interrupted by user")
            break
        except Exception as e:
            log.exception("video failed, continuing: %s", e)

    log.info("batch complete")


def reset_dataset() -> None:
    """DANGEROUS: wipe the calibration dataset. Keeps backups."""
    if DATASET_JSON.exists():
        backup = DATASET_JSON.with_suffix(".json.wiped")
        DATASET_JSON.rename(backup)
        log.warning("Dataset wiped. Backup saved to %s", backup)
    if DATASET_XLSX.exists():
        DATASET_XLSX.unlink()


# ── CLI entry point ──────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="TRIBE v2 Neural Content Optimization Calibrator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--url", help="Analyze a single YouTube URL")
    group.add_argument("--batch", help="Analyze a batch of URLs from a file")
    group.add_argument("--decision-card", action="store_true", help="Compute GO/REFINE/PIVOT verdict")
    group.add_argument("--reset-dataset", action="store_true", help="Wipe the dataset (keeps backup)")

    parser.add_argument(
        "--category",
        default="uncategorized",
        choices=["viral", "competitor", "anti_example", "uncategorized"],
        help="Category for --url mode",
    )
    return parser


def main() -> int:
    NEURAL_CAL_DIR.mkdir(parents=True, exist_ok=True)
    VIDEOS_DIR.mkdir(parents=True, exist_ok=True)
    TRIBE_DIR.mkdir(parents=True, exist_ok=True)

    args = build_parser().parse_args()

    log.info("TRIBE v2 Calibrator starting - mode=%s", _mode_name(args))

    try:
        if args.reset_dataset:
            reset_dataset()
            return 0
        if args.decision_card:
            result = decision_card()
            log.info("Decision: %s", result.get("verdict"))
            return 0
        if args.url:
            process_single_url(args.url, args.category)
            return 0
        if args.batch:
            process_batch(Path(args.batch))
            return 0
    except NotImplementedError as e:
        log.warning("not implemented yet: %s", e)
        return 0
    except Exception as e:
        log.exception("fatal error: %s", e)
        return 1

    return 0


def _mode_name(args: argparse.Namespace) -> str:
    if args.url:
        return "single-url"
    if args.batch:
        return "batch"
    if args.decision_card:
        return "decision-card"
    if args.reset_dataset:
        return "reset-dataset"
    return "unknown"


if __name__ == "__main__":
    sys.exit(main())
